from datetime import date, datetime, timedelta
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import logout
from django.utils import timezone
from django.db.models import Q, Count
from django.urls import reverse
from django.contrib.auth.models import User
import matplotlib
matplotlib.use('Agg')  # Para evitar problemas com GUI
import matplotlib.pyplot as plt
import io
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from accounts.models import Professor
from cadastros.models import HorarioAula, Disciplina, Turma
from permuta.models import Permuta, Reposicao, Notificacao
from permuta.forms import PermutaSolicitacaoForm, ReposicaoForm
from permuta.utils import (
    notificar_nova_permuta,
    notificar_confirmacao_permuta,
    notificar_cancelamento_permuta
)


# ============================================================================
# PÁGINAS PÚBLICAS
# ============================================================================

def home(request):
    """
    Página inicial do sistema.
    - Se usuário não estiver logado: mostra página institucional
    - Se usuário estiver logado: redireciona para dashboard específico
    """
    usuario = request.user
    
    if usuario.is_authenticated:
        # Usuário logado - redireciona para dashboard específico
        try:
            professor = usuario.professor
            
            if usuario.is_staff:
                # Usuário é admin/coordenação - redireciona para dashboard admin
                return redirect('admin_dashboard')
            else:
                # Usuário é professor normal - redireciona para dashboard professor
                return redirect('professor_dashboard')
                
        except Professor.DoesNotExist:
            # Usuário logado mas não é professor (só admin)
            if usuario.is_staff:
                return redirect('admin_dashboard')
            else:
                messages.warning(request, "Seu usuário não está vinculado a um perfil de professor.")
                return render(request, "sem_professor.html", {"usuario": usuario})
    else:
        # Usuário não logado - mostra página institucional
        return render(request, "home.html")


def meu_logout(request):
    """
    Faz logout do usuário e redireciona para a home.
    """
    logout(request)
    return redirect("home")


# ============================================================================
# DASHBOARDS SEPARADOS
# ============================================================================

@login_required
def professor_dashboard(request):
    """
    Dashboard específico para professores
    """
    usuario = request.user
    
    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(request, "Acesso restrito a professores.")
        return redirect("home")
    
    # Estatísticas do professor
    total_horarios = HorarioAula.objects.filter(professor=professor).count()
    total_permutas = Permuta.objects.filter(professor_solicitante=professor).count()
    permutas_pendentes = Permuta.objects.filter(
        professor_solicitante=professor, 
        status='PENDENTE'
    ).count()
    permutas_aprovadas = Permuta.objects.filter(
        professor_solicitante=professor, 
        status='APROVADA'
    ).count()
    permutas_canceladas = Permuta.objects.filter(
        professor_solicitante=professor, 
        status='CANCELADA'
    ).count()
    
    # Próximas permutas (aulas futuras)
    proximas_permutas = Permuta.objects.filter(
        professor_solicitante=professor,
        data_aula__gte=timezone.now().date()
    ).order_by('data_aula')[:5]
    
    # Permutas como substituto
    permutas_substituto = Permuta.objects.filter(
        professor_substituto=professor
    ).count()
    
    permutas_substituto_pendentes = Permuta.objects.filter(
        professor_substituto=professor,
        status='PENDENTE'
    ).count()
    
    contexto = {
        'usuario': usuario,
        'professor': professor,
        'total_horarios': total_horarios,
        'total_permutas': total_permutas,
        'permutas_pendentes': permutas_pendentes,
        'permutas_aprovadas': permutas_aprovadas,
        'permutas_canceladas': permutas_canceladas,
        'permutas_substituto': permutas_substituto,
        'permutas_substituto_pendentes': permutas_substituto_pendentes,
        'proximas_permutas': proximas_permutas,
    }
    return render(request, "professor/dashboard.html", contexto)


@login_required
def admin_dashboard(request):
    """
    Dashboard específico para administradores/coordenação
    """
    usuario = request.user
    
    if not usuario.is_staff:
        messages.error(request, "Acesso restrito à coordenação.")
        return redirect("home")
    
    # Estatísticas gerais
    total_professores = Professor.objects.count()
    total_turmas = Turma.objects.count()
    total_disciplinas = Disciplina.objects.count()
    total_horarios = HorarioAula.objects.count()
    
    total_permutas = Permuta.objects.count()
    permutas_aprovadas = Permuta.objects.filter(status='APROVADA').count()
    permutas_pendentes = Permuta.objects.filter(status='PENDENTE').count()
    permutas_canceladas = Permuta.objects.filter(status='CANCELADA').count()
    
    # Permutas sem reposição
    todas_permutas = Permuta.objects.all()
    permutas_sem_reposicao = [p for p in todas_permutas if not p.tem_reposicao()]
    
    # Top 5 professores que mais solicitam permutas
    top_professores = Professor.objects.annotate(
        total_permutas=Count('permutas_solicitadas')
    ).order_by('-total_permutas')[:5]
    
    # Top 5 disciplinas mais permutadas
    top_disciplinas = Disciplina.objects.annotate(
        total_permutas=Count('horarioaula__permuta')
    ).order_by('-total_permutas')[:5]
    
    # Permutas por mês (últimos 6 meses)
    hoje = timezone.now().date()
    permutas_por_mes = []
    for i in range(5, -1, -1):
        mes = hoje - timedelta(days=30*i)
        count = Permuta.objects.filter(
            data_solicitacao__month=mes.month,
            data_solicitacao__year=mes.year
        ).count()
        permutas_por_mes.append({
            'mes': mes.strftime('%b/%Y'),
            'quantidade': count
        })
    
    contexto = {
        'usuario': usuario,
        'total_professores': total_professores,
        'total_turmas': total_turmas,
        'total_disciplinas': total_disciplinas,
        'total_horarios': total_horarios,
        'total_permutas': total_permutas,
        'permutas_aprovadas': permutas_aprovadas,
        'permutas_pendentes': permutas_pendentes,
        'permutas_canceladas': permutas_canceladas,
        'permutas_sem_reposicao': permutas_sem_reposicao[:5],
        'top_professores': top_professores,
        'top_disciplinas': top_disciplinas,
        'permutas_por_mes': permutas_por_mes,
    }
    return render(request, "admin/dashboard.html", contexto)


# ============================================================================
# PROFESSOR - HORÁRIOS
# ============================================================================

@login_required
def meus_horarios(request):
    """
    Lista os horários de aula do professor logado.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor. "
            "Peça ao administrador para criar o seu cadastro de Professor."
        )
        return render(request, "sem_professor.html", {"usuario": usuario})

    horarios = HorarioAula.objects.filter(
        professor=professor
    ).order_by("dia_semana", "hora_inicio")

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "horarios": horarios,
    }
    return render(request, "professor/meus_horarios.html", contexto)


@login_required
def calendario_horarios(request):
    """
    Visualização em calendário dos horários do professor
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor."
        )
        return redirect("home")

    contexto = {
        "usuario": usuario,
        "professor": professor,
    }
    return render(request, "professor/calendario_horarios.html", contexto)


@login_required
def api_eventos_calendario(request):
    """
    API para fornecer eventos ao calendário (FullCalendar)
    """
    usuario = request.user
    start = request.GET.get('start')
    end = request.GET.get('end')
    
    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        return JsonResponse([], safe=False)
    
    eventos = []
    
    # Buscar horários regulares do professor
    horarios = HorarioAula.objects.filter(professor=professor)
    
    # Mapear dias da semana
    dias_semana = {
        2: 'monday', 3: 'tuesday', 4: 'wednesday',
        5: 'thursday', 6: 'friday', 7: 'saturday'
    }
    
    for horario in horarios:
        dia_semana = dias_semana.get(horario.dia_semana)
        if dia_semana:
            eventos.append({
                'title': f"{horario.disciplina.nome} - {horario.turma.codigo_turma}",
                'daysOfWeek': [dia_semana],
                'startTime': str(horario.hora_inicio),
                'endTime': str(horario.hora_fim),
                'color': '#28a745',
                'url': reverse('solicitar_permuta', args=[horario.id]),
                'extendedProps': {
                    'tipo': 'horario_fixo',
                    'disciplina': horario.disciplina.nome,
                    'turma': horario.turma.codigo_turma,
                    'horario_id': horario.id
                }
            })
    
    # Buscar permutas do professor
    permutas = Permuta.objects.filter(
        Q(professor_solicitante=professor) | Q(professor_substituto=professor)
    )
    
    for permuta in permutas:
        cor = '#dc3545' if permuta.status == 'CANCELADA' else '#ffc107' if permuta.status == 'PENDENTE' else '#28a745'
        
        eventos.append({
            'title': f"Permuta #{permuta.id} - {permuta.get_status_display()}",
            'start': permuta.data_aula.strftime('%Y-%m-%d'),
            'allDay': True,
            'color': cor,
            'url': reverse('detalhe_permuta', args=[permuta.id]),
            'extendedProps': {
                'tipo': 'permuta',
                'status': permuta.status,
                'solicitante': permuta.professor_solicitante.nome,
                'substituto': permuta.professor_substituto.nome
            }
        })
        
        if permuta.tem_reposicao():
            eventos.append({
                'title': f"Reposição #{permuta.id}",
                'start': permuta.reposicao.data_reposicao.strftime('%Y-%m-%d'),
                'allDay': True,
                'color': '#17a2b8',
                'url': reverse('detalhe_permuta', args=[permuta.id]),
                'extendedProps': {
                    'tipo': 'reposicao',
                    'permuta_id': permuta.id
                }
            })
    
    return JsonResponse(eventos, safe=False)


# ============================================================================
# PROFESSOR - PERMUTAS
# ============================================================================

@login_required
def solicitar_permuta(request, horario_id):
    """
    Permite ao professor logado solicitar permuta para um horário específico.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor. "
            "Peça ao administrador para criar o seu cadastro de Professor."
        )
        return redirect("home")

    horario = get_object_or_404(HorarioAula, id=horario_id, professor=professor)

    if request.method == "POST":
        form = PermutaSolicitacaoForm(request.POST, professor_solicitante=professor)
        if form.is_valid():
            permuta = form.save(commit=False)
            permuta.professor_solicitante = professor
            permuta.horario = horario
            permuta.status = "PENDENTE"
            permuta.save()

            # Notificação no sistema
            substituto_user = permuta.professor_substituto.user
            Notificacao.objects.create(
                usuario=substituto_user,
                mensagem=(
                    f"Você foi indicado(a) como professor(a) substituto(a) "
                    f"na permuta #{permuta.id} do professor {permuta.professor_solicitante.nome}."
                ),
                link=reverse("detalhe_permuta", args=[permuta.id]),
            )

            # Notificação por email
            notificar_nova_permuta(permuta, request)

            messages.success(
                request,
                "Solicitação de permuta registrada com sucesso! Agora registre a data da reposição."
            )
            return redirect("registrar_reposicao", permuta_id=permuta.id)
    else:
        form = PermutaSolicitacaoForm(
            professor_solicitante=professor,
            initial={"data_aula": date.today()},
        )

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "horario": horario,
        "form": form,
    }
    return render(request, "professor/solicitar_permuta.html", contexto)


@login_required
def minhas_permutas(request):
    """
    Lista todas as solicitações de permuta do professor logado.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor. "
            "Peça ao administrador para criar o seu cadastro de Professor."
        )
        return redirect("home")

    permutas = Permuta.objects.filter(
        professor_solicitante=professor
    ).order_by("-data_solicitacao")

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "permutas": permutas,
    }
    return render(request, "professor/minhas_permutas.html", contexto)


@login_required
def detalhe_permuta(request, permuta_id):
    """
    Mostra os detalhes de uma permuta específica.
    """
    usuario = request.user
    professor = None

    if usuario.is_staff:
        permuta = get_object_or_404(Permuta, id=permuta_id)
    else:
        try:
            professor = usuario.professor
        except Professor.DoesNotExist:
            messages.error(
                request,
                "Seu usuário ainda não está vinculado a um cadastro de Professor."
            )
            return redirect("home")

        permuta = get_object_or_404(
            Permuta,
            Q(id=permuta_id) & (
                Q(professor_solicitante=professor) |
                Q(professor_substituto=professor)
            )
        )

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "permuta": permuta,
    }
    return render(request, "professor/detalhe_permuta.html", contexto)


@login_required
def cancelar_permuta(request, permuta_id):
    """
    Permite ao professor solicitante cancelar uma permuta.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor."
        )
        return redirect("home")

    permuta = get_object_or_404(Permuta, id=permuta_id, professor_solicitante=professor)

    if permuta.status == "CANCELADA":
        messages.error(
            request,
            "Esta permuta já foi cancelada anteriormente."
        )
        return redirect("minhas_permutas")

    if request.method == "POST":
        permuta.status = "CANCELADA"
        permuta.data_decisao = timezone.now()
        permuta.usuario_decisor = usuario
        permuta.save()

        # Notificação no sistema
        substituto_user = permuta.professor_substituto.user
        Notificacao.objects.create(
            usuario=substituto_user,
            mensagem=(
                f"A permuta #{permuta.id} com o professor "
                f"{permuta.professor_solicitante.nome} foi CANCELADA pelo solicitante."
            ),
            link=reverse("detalhe_permuta", args=[permuta.id]),
        )

        # Notificações para coordenadores
        coordenadores = User.objects.filter(is_staff=True)
        for coord in coordenadores:
            Notificacao.objects.create(
                usuario=coord,
                mensagem=(
                    f"A permuta #{permuta.id} do professor {permuta.professor_solicitante.nome} "
                    f"com o professor {permuta.professor_substituto.nome} foi CANCELADA."
                ),
                link=reverse("detalhe_permuta", args=[permuta.id]),
            )

        # Notificação por email
        notificar_cancelamento_permuta(permuta, request)

        messages.success(
            request,
            "Permuta cancelada com sucesso."
        )
        return redirect("minhas_permutas")

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "permuta": permuta,
    }
    return render(request, "professor/cancelar_permuta.html", contexto)


@login_required
def permutas_como_substituto(request):
    """
    Lista as permutas em que o professor logado é o professor_substituto.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor."
        )
        return redirect("home")

    permutas = Permuta.objects.filter(
        professor_substituto=professor
    ).order_by("-data_solicitacao")

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "permutas": permutas,
    }
    return render(request, "professor/permutas_como_substituto.html", contexto)


@login_required
def confirmar_permuta_substituto(request, permuta_id):
    """
    Permite ao professor substituto confirmar uma permuta.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor."
        )
        return redirect("home")

    permuta = get_object_or_404(
        Permuta,
        id=permuta_id,
        professor_substituto=professor
    )

    if permuta.status == "CANCELADA":
        messages.error(
            request,
            "Esta permuta foi cancelada e não pode mais ser confirmada."
        )
        return redirect("permutas_como_substituto")

    if permuta.status == "APROVADA":
        messages.info(
            request,
            "Esta permuta já foi aprovada anteriormente."
        )
        return redirect("permutas_como_substituto")

    if not permuta.tem_reposicao():
        messages.error(
            request,
            "Não é possível confirmar esta permuta, pois o professor solicitante ainda não registrou a reposição."
        )
        return redirect("permutas_como_substituto")

    permuta.status = "APROVADA"
    permuta.data_decisao = timezone.now()
    permuta.usuario_decisor = usuario
    permuta.save()

    # Notificações para coordenadores
    coordenadores = User.objects.filter(is_staff=True)
    for coord in coordenadores:
        Notificacao.objects.create(
            usuario=coord,
            mensagem=(
                f"A permuta #{permuta.id} entre "
                f"{permuta.professor_solicitante.nome} e "
                f"{permuta.professor_substituto.nome} foi APROVADA pelo professor substituto."
            ),
            link=reverse("detalhe_permuta", args=[permuta.id]),
        )

    # Notificação por email
    notificar_confirmacao_permuta(permuta, request)

    messages.success(
        request,
        "Permuta aprovada com sucesso."
    )
    return redirect("permutas_como_substituto")


@login_required
def registrar_reposicao(request, permuta_id):
    """
    Permite ao professor solicitante registrar a reposição de uma permuta.
    """
    usuario = request.user

    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        messages.error(
            request,
            "Seu usuário ainda não está vinculado a um cadastro de professor."
        )
        return redirect("home")

    permuta = get_object_or_404(
        Permuta,
        id=permuta_id,
        professor_solicitante=professor,
    )

    if hasattr(permuta, "reposicao"):
        messages.error(
            request,
            "Esta permuta já possui uma reposição registrada."
        )
        return redirect("minhas_permutas")

    if request.method == "POST":
        form = ReposicaoForm(request.POST)
        if form.is_valid():
            reposicao = form.save(commit=False)
            reposicao.permuta = permuta
            reposicao.save()

            # Notificação para o substituto
            Notificacao.objects.create(
                usuario=permuta.professor_substituto.user,
                mensagem=(
                    f"O professor {professor.nome} registrou a reposição para a permuta #{permuta.id}. "
                    f"Agora você pode confirmar a permuta."
                ),
                link=reverse("confirmar_permuta_substituto", args=[permuta.id]),
            )

            messages.success(
                request,
                "Reposição registrada com sucesso! Agora o professor substituto deve confirmar a permuta."
            )
            return redirect("detalhe_permuta", permuta_id=permuta.id)
    else:
        form = ReposicaoForm()

    contexto = {
        "usuario": usuario,
        "professor": professor,
        "permuta": permuta,
        "form": form,
    }
    return render(request, "professor/registrar_reposicao.html", contexto)


# ============================================================================
# COORDENAÇÃO / ADMIN
# ============================================================================

@login_required
def permutas_pendentes(request):
    """
    Visão da coordenação: lista permutas solicitadas que ainda não têm reposição.
    """
    usuario = request.user

    if not usuario.is_staff:
        messages.error(request, "Você não tem permissão para acessar esta página.")
        return redirect("home")

    permutas = Permuta.objects.all().order_by("-data_solicitacao")
    permutas_sem_reposicao = [p for p in permutas if not p.tem_reposicao()]

    contexto = {
        "usuario": usuario,
        "permutas": permutas_sem_reposicao,
    }
    return render(request, "coordenacao/permutas_pendentes.html", contexto)


@login_required
def dashboard_estatisticas(request):
    """
    Dashboard com gráficos e estatísticas das permutas
    Apenas para staff/coordenação
    """
    usuario = request.user
    
    if not usuario.is_staff:
        messages.error(request, "Acesso restrito à coordenação.")
        return redirect("home")
    
    hoje = timezone.now().date()
    
    # Estatísticas gerais
    total_permutas = Permuta.objects.count()
    permutas_aprovadas = Permuta.objects.filter(status='APROVADA').count()
    permutas_pendentes = Permuta.objects.filter(status='PENDENTE').count()
    permutas_canceladas = Permuta.objects.filter(status='CANCELADA').count()
    
    # Permutas por mês (últimos 6 meses)
    meses = []
    dados_mensais = []
    
    for i in range(5, -1, -1):
        data = hoje - timedelta(days=30*i)
        meses.append(data.strftime('%b/%Y'))
        
        mes_inicio = hoje - timedelta(days=30*(i+1))
        mes_fim = hoje - timedelta(days=30*i)
        count = Permuta.objects.filter(
            data_solicitacao__date__gte=mes_inicio,
            data_solicitacao__date__lt=mes_fim
        ).count()
        dados_mensais.append(count)
    
    # Gráfico de Pizza - Status das Permutas
    fig1, ax1 = plt.subplots(figsize=(6,4))
    status_counts = [permutas_aprovadas, permutas_pendentes, permutas_canceladas]
    status_labels = ['Aprovadas', 'Pendentes', 'Canceladas']
    colors = ['#28a745', '#ffc107', '#dc3545']
    
    if sum(status_counts) > 0:
        ax1.pie(status_counts, labels=status_labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax1.axis('equal')
        ax1.set_title('Distribuição por Status', fontsize=14, fontweight='bold')
    else:
        ax1.text(0.5, 0.5, 'Sem dados', ha='center', va='center')
        ax1.set_xlim(-1, 1)
        ax1.set_ylim(-1, 1)
    
    buffer1 = io.BytesIO()
    plt.savefig(buffer1, format='png', bbox_inches='tight')
    buffer1.seek(0)
    grafico_pizza = base64.b64encode(buffer1.getvalue()).decode()
    plt.close(fig1)
    
    # Gráfico de Barras - Permutas por Mês
    fig2, ax2 = plt.subplots(figsize=(8,4))
    bars = ax2.bar(meses, dados_mensais, color='#28a745', alpha=0.7)
    ax2.set_xlabel('Mês')
    ax2.set_ylabel('Quantidade')
    ax2.set_title('Permutas por Mês (últimos 6 meses)', fontsize=14, fontweight='bold')
    ax2.tick_params(axis='x', rotation=45)
    
    for bar in bars:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}', ha='center', va='bottom')
    
    buffer2 = io.BytesIO()
    plt.savefig(buffer2, format='png', bbox_inches='tight')
    buffer2.seek(0)
    grafico_barras = base64.b64encode(buffer2.getvalue()).decode()
    plt.close(fig2)
    
    # Top 5 professores que mais solicitam permutas
    top_professores = Professor.objects.annotate(
        total_permutas=Count('permutas_solicitadas')
    ).order_by('-total_permutas')[:5]
    
    # Top 5 disciplinas mais permutadas
    top_disciplinas = Disciplina.objects.annotate(
        total_permutas=Count('horarioaula__permuta')
    ).order_by('-total_permutas')[:5]
    
    # Permutas por dia da semana
    dias_semana = {
        2: 'Segunda', 3: 'Terça', 4: 'Quarta',
        5: 'Quinta', 6: 'Sexta', 7: 'Sábado'
    }
    permutas_por_dia = []
    for dia_num, dia_nome in dias_semana.items():
        count = Permuta.objects.filter(horario__dia_semana=dia_num).count()
        permutas_por_dia.append({'dia': dia_nome, 'quantidade': count})
    
    contexto = {
        'usuario': usuario,
        'total_permutas': total_permutas,
        'permutas_aprovadas': permutas_aprovadas,
        'permutas_pendentes': permutas_pendentes,
        'permutas_canceladas': permutas_canceladas,
        'grafico_pizza': grafico_pizza,
        'grafico_barras': grafico_barras,
        'top_professores': top_professores,
        'top_disciplinas': top_disciplinas,
        'permutas_por_dia': permutas_por_dia,
    }
    return render(request, "coordenacao/dashboard_estatisticas.html", contexto)


@login_required
def relatorio_permutas_excel(request):
    """
    Gera relatório em Excel com todas as permutas
    """
    usuario = request.user
    
    if not usuario.is_staff:
        messages.error(request, "Acesso restrito à coordenação.")
        return redirect("home")
    
    # Definir a variável hoje
    hoje = timezone.now().date()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Permutas"
    
    # Cabeçalhos
    headers = [
        'ID', 'Data Solicitação', 'Data Aula', 'Solicitante', 'Substituto',
        'Turma', 'Disciplina', 'Status', 'Data Decisão', 'Tem Reposição',
        'Data Reposição', 'Motivo'
    ]
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Dados
    permutas = Permuta.objects.all().order_by('-data_solicitacao')
    
    for row_num, permuta in enumerate(permutas, 2):
        ws.cell(row=row_num, column=1).value = permuta.id
        ws.cell(row=row_num, column=2).value = permuta.data_solicitacao.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_num, column=3).value = permuta.data_aula.strftime('%d/%m/%Y')
        ws.cell(row=row_num, column=4).value = permuta.professor_solicitante.nome
        ws.cell(row=row_num, column=5).value = permuta.professor_substituto.nome
        ws.cell(row=row_num, column=6).value = permuta.horario.turma.codigo_turma
        ws.cell(row=row_num, column=7).value = permuta.horario.disciplina.nome
        ws.cell(row=row_num, column=8).value = permuta.get_status_display()
        ws.cell(row=row_num, column=9).value = permuta.data_decisao.strftime('%d/%m/%Y %H:%M') if permuta.data_decisao else ''
        ws.cell(row=row_num, column=10).value = 'Sim' if permuta.tem_reposicao() else 'Não'
        
        if permuta.tem_reposicao():
            ws.cell(row=row_num, column=11).value = permuta.reposicao.data_reposicao.strftime('%d/%m/%Y')
        else:
            ws.cell(row=row_num, column=11).value = ''
            
        ws.cell(row=row_num, column=12).value = permuta.motivo
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=relatorio_permutas_{hoje.strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def relatorio_permutas_pdf(request):
    """
    Gera relatório em PDF com todas as permutas
    """
    usuario = request.user
    
    if not usuario.is_staff:
        messages.error(request, "Acesso restrito à coordenação.")
        return redirect("home")
    
    # Definir a variável hoje
    hoje = timezone.now().date()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=relatorio_permutas_{hoje.strftime("%Y%m%d")}.pdf'
    
    p = canvas.Canvas(response, pagesize=A4)
    largura, altura = A4
    
    # Cabeçalho
    p.setFillColor(colors.HexColor('#28a745'))
    p.rect(0, altura-80, largura, 80, fill=1)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 20)
    p.drawString(50, altura-45, "Relatório de Permutas")
    p.setFont("Helvetica", 10)
    p.drawString(50, altura-65, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    y = altura - 120
    x = 50
    
    permutas = Permuta.objects.all().order_by('-data_solicitacao')[:50]  # Limitar a 50 para PDF
    
    for permuta in permutas:
        if y < 50:  # Nova página
            p.showPage()
            y = altura - 50
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Continuação...")
            y -= 30
        
        p.setFont("Helvetica-Bold", 11)
        p.drawString(x, y, f"Permuta #{permuta.id}")
        y -= 15
        
        p.setFont("Helvetica", 9)
        p.drawString(x+20, y, f"Solicitante: {permuta.professor_solicitante.nome}")
        y -= 12
        p.drawString(x+20, y, f"Substituto: {permuta.professor_substituto.nome}")
        y -= 12
        p.drawString(x+20, y, f"Data Aula: {permuta.data_aula.strftime('%d/%m/%Y')}")
        y -= 12
        p.drawString(x+20, y, f"Turma: {permuta.horario.turma.codigo_turma} - {permuta.horario.disciplina.nome}")
        y -= 12
        p.drawString(x+20, y, f"Status: {permuta.get_status_display()}")
        y -= 20
        
        # Linha separadora
        p.setStrokeColor(colors.lightgrey)
        p.line(x, y, largura-50, y)
        y -= 15
    
    p.save()
    return response


# ============================================================================
# API REST
# ============================================================================

@login_required
def api_permutas(request):
    """
    API REST para listar permutas em formato JSON
    """
    usuario = request.user
    
    if usuario.is_staff:
        permutas = Permuta.objects.all()
    else:
        try:
            professor = usuario.professor
            permutas = Permuta.objects.filter(
                Q(professor_solicitante=professor) | Q(professor_substituto=professor)
            )
        except Professor.DoesNotExist:
            return JsonResponse({'error': 'Professor não encontrado'}, status=404)
    
    data = []
    for permuta in permutas.order_by('-data_solicitacao')[:100]:
        data.append({
            'id': permuta.id,
            'data_solicitacao': permuta.data_solicitacao.strftime('%Y-%m-%d %H:%M:%S'),
            'data_aula': permuta.data_aula.strftime('%Y-%m-%d'),
            'solicitante': {
                'id': permuta.professor_solicitante.id,
                'nome': permuta.professor_solicitante.nome,
                'siape': permuta.professor_solicitante.matricula_siape,
            },
            'substituto': {
                'id': permuta.professor_substituto.id,
                'nome': permuta.professor_substituto.nome,
                'siape': permuta.professor_substituto.matricula_siape,
            },
            'turma': {
                'codigo': permuta.horario.turma.codigo_turma,
                'descricao': permuta.horario.turma.descricao,
            },
            'disciplina': {
                'nome': permuta.horario.disciplina.nome,
            },
            'horario': {
                'dia_semana': permuta.horario.get_dia_semana_display(),
                'hora_inicio': str(permuta.horario.hora_inicio),
                'hora_fim': str(permuta.horario.hora_fim),
            },
            'status': permuta.status,
            'status_display': permuta.get_status_display(),
            'motivo': permuta.motivo,
            'tem_reposicao': permuta.tem_reposicao(),
            'data_decisao': permuta.data_decisao.strftime('%Y-%m-%d %H:%M:%S') if permuta.data_decisao else None,
            'url_detalhes': reverse('detalhe_permuta', args=[permuta.id]),
            'url_pdf': reverse('comprovante_permuta_pdf', args=[permuta.id]),
        })
    
    return JsonResponse({'count': len(data), 'results': data}, safe=False)


@login_required
def api_permuta_detalhe(request, permuta_id):
    """
    API REST para detalhes de uma permuta específica
    """
    usuario = request.user
    
    try:
        if usuario.is_staff:
            permuta = Permuta.objects.get(id=permuta_id)
        else:
            professor = usuario.professor
            permuta = Permuta.objects.get(
                Q(id=permuta_id) & (
                    Q(professor_solicitante=professor) | Q(professor_substituto=professor)
                )
            )
    except Professor.DoesNotExist:
        return JsonResponse({'error': 'Professor não encontrado'}, status=404)
    except Permuta.DoesNotExist:
        return JsonResponse({'error': 'Permuta não encontrada'}, status=404)
    
    data = {
        'id': permuta.id,
        'data_solicitacao': permuta.data_solicitacao.strftime('%Y-%m-%d %H:%M:%S'),
        'data_aula': permuta.data_aula.strftime('%Y-%m-%d'),
        'solicitante': {
            'id': permuta.professor_solicitante.id,
            'nome': permuta.professor_solicitante.nome,
            'siape': permuta.professor_solicitante.matricula_siape,
            'email': permuta.professor_solicitante.email,
        },
        'substituto': {
            'id': permuta.professor_substituto.id,
            'nome': permuta.professor_substituto.nome,
            'siape': permuta.professor_substituto.matricula_siape,
            'email': permuta.professor_substituto.email,
        },
        'turma': {
            'codigo': permuta.horario.turma.codigo_turma,
            'descricao': permuta.horario.turma.descricao,
        },
        'disciplina': {
            'id': permuta.horario.disciplina.id,
            'nome': permuta.horario.disciplina.nome,
        },
        'horario': {
            'id': permuta.horario.id,
            'dia_semana': permuta.horario.dia_semana,
            'dia_semana_display': permuta.horario.get_dia_semana_display(),
            'hora_inicio': str(permuta.horario.hora_inicio),
            'hora_fim': str(permuta.horario.hora_fim),
        },
        'status': permuta.status,
        'status_display': permuta.get_status_display(),
        'motivo': permuta.motivo,
        'tem_reposicao': permuta.tem_reposicao(),
        'data_decisao': permuta.data_decisao.strftime('%Y-%m-%d %H:%M:%S') if permuta.data_decisao else None,
        'usuario_decisor': permuta.usuario_decisor.username if permuta.usuario_decisor else None,
        'reposicao': {
            'data_reposicao': permuta.reposicao.data_reposicao.strftime('%Y-%m-%d') if permuta.tem_reposicao() else None,
            'observacao': permuta.reposicao.observacao if permuta.tem_reposicao() else None,
        } if permuta.tem_reposicao() else None,
        'urls': {
            'detalhes': reverse('detalhe_permuta', args=[permuta.id]),
            'pdf': reverse('comprovante_permuta_pdf', args=[permuta.id]),
            'cancelar': reverse('cancelar_permuta', args=[permuta.id]) if permuta.status != 'CANCELADA' else None,
            'confirmar': reverse('confirmar_permuta_substituto', args=[permuta.id]) if permuta.status == 'PENDENTE' and permuta.tem_reposicao() else None,
        }
    }
    
    return JsonResponse(data)


@login_required
def api_estatisticas(request):
    """
    API REST para estatísticas das permutas
    """
    usuario = request.user
    
    if not usuario.is_staff:
        return JsonResponse({'error': 'Acesso restrito'}, status=403)
    
    hoje = timezone.now().date()
    
    # Estatísticas gerais
    total = Permuta.objects.count()
    aprovadas = Permuta.objects.filter(status='APROVADA').count()
    pendentes = Permuta.objects.filter(status='PENDENTE').count()
    canceladas = Permuta.objects.filter(status='CANCELADA').count()
    
    # Permutas por mês
    permutas_por_mes = []
    for i in range(5, -1, -1):
        mes = hoje - timedelta(days=30*i)
        count = Permuta.objects.filter(
            data_solicitacao__month=mes.month,
            data_solicitacao__year=mes.year
        ).count()
        permutas_por_mes.append({
            'mes': mes.strftime('%Y-%m'),
            'quantidade': count
        })
    
    data = {
        'total_permutas': total,
        'aprovadas': aprovadas,
        'pendentes': pendentes,
        'canceladas': canceladas,
        'taxa_aprovacao': round((aprovadas / total * 100) if total > 0 else 0, 2),
        'permutas_por_mes': permutas_por_mes,
        'timestamp': datetime.now().isoformat(),
    }
    
    return JsonResponse(data)


# ============================================================================
# NOTIFICAÇÕES
# ============================================================================

@login_required
def ler_notificacao(request, notificacao_id):
    """
    Marca a notificação como lida e redireciona para o link associado.
    """
    notificacao = get_object_or_404(
        Notificacao,
        id=notificacao_id,
        usuario=request.user
    )

    notificacao.lida = True
    notificacao.save()

    if notificacao.link:
        return redirect(notificacao.link)
    return redirect("home")


@login_required
def api_notificacoes_nao_lidas(request):
    """
    API REST para listar notificações não lidas
    """
    notificacoes = Notificacao.objects.filter(
        usuario=request.user,
        lida=False
    ).order_by('-criada_em')[:10]
    
    data = []
    for notif in notificacoes:
        data.append({
            'id': notif.id,
            'mensagem': notif.mensagem,
            'criada_em': notif.criada_em.strftime('%d/%m/%Y %H:%M'),
            'link': notif.link,
            'url_ler': reverse('ler_notificacao', args=[notif.id]),
        })
    
    return JsonResponse({'count': len(data), 'notificacoes': data})


# ============================================================================
# COMPROVANTES E RELATÓRIOS
# ============================================================================

@login_required
def comprovante_permuta_pdf(request, permuta_id):
    """
    Gera um comprovante profissional em PDF com os dados da permuta.
    """
    usuario = request.user

    professor = None
    try:
        professor = usuario.professor
    except Professor.DoesNotExist:
        professor = None

    permuta = get_object_or_404(Permuta, id=permuta_id)

    if not (
        (professor and permuta.professor_solicitante == professor)
        or (professor and permuta.professor_substituto == professor)
        or usuario.is_staff
    ):
        messages.error(
            request,
            "Você não tem permissão para acessar o comprovante desta permuta."
        )
        return redirect("home")

    response = HttpResponse(content_type="application/pdf")
    filename = f"comprovante_permuta_{permuta.id}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Create the PDF object
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Cabeçalho com gradiente
    p.setFillColor(colors.HexColor('#28a745'))
    p.rect(0, height-100, width, 100, fill=1)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 24)
    p.drawString(50, height-50, "Sistema de Permuta de Aulas")
    p.setFont("Helvetica", 12)
    p.drawString(50, height-75, "Comprovante de Permuta")

    # Data de emissão
    p.setFont("Helvetica", 10)
    p.drawRightString(width-50, height-40, f"Emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # Código da permuta
    p.setFillColor(colors.HexColor('#dc3545'))
    p.setFont("Helvetica-Bold", 20)
    p.drawString(50, height-150, f"Permuta #{permuta.id}")
    
    # Status
    status_colors = {
        'PENDENTE': '#ffc107',
        'APROVADA': '#28a745',
        'CANCELADA': '#dc3545',
    }
    status_color = status_colors.get(permuta.status, '#6c757d')
    p.setFillColor(colors.HexColor(status_color))
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width-50, height-150, f"Status: {permuta.get_status_display()}")

    y = height - 200

    # Função para desenhar seções
    def draw_section(title, y_pos):
        p.setFillColor(colors.HexColor('#28a745'))
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_pos, title)
        p.setFillColor(colors.HexColor('#dc3545'))
        p.line(50, y_pos-5, width-50, y_pos-5)
        return y_pos - 30

    # Seção: Professor Solicitante
    y = draw_section("Professor Solicitante", y)
    p.setFillColor(colors.black)
    p.setFont("Helvetica", 11)
    p.drawString(70, y, f"Nome: {permuta.professor_solicitante.nome}")
    y -= 20
    p.drawString(70, y, f"Matrícula SIAPE: {permuta.professor_solicitante.matricula_siape}")
    y -= 30

    # Seção: Professor Substituto
    y = draw_section("Professor Substituto", y)
    p.setFont("Helvetica", 11)
    p.drawString(70, y, f"Nome: {permuta.professor_substituto.nome}")
    y -= 20
    p.drawString(70, y, f"Matrícula SIAPE: {permuta.professor_substituto.matricula_siape}")
    y -= 30

    # Seção: Dados da Aula
    y = draw_section("Dados da Aula", y)
    p.setFont("Helvetica", 11)
    p.drawString(70, y, f"Turma: {permuta.horario.turma.codigo_turma}")
    y -= 20
    p.drawString(70, y, f"Disciplina: {permuta.horario.disciplina.nome}")
    y -= 20
    p.drawString(70, y, f"Dia da semana: {permuta.horario.get_dia_semana_display()}")
    y -= 20
    p.drawString(70, y, f"Horário: {permuta.horario.hora_inicio} - {permuta.horario.hora_fim}")
    y -= 20
    p.drawString(70, y, f"Data da aula permutada: {permuta.data_aula.strftime('%d/%m/%Y')}")
    y -= 30

    # Seção: Datas
    y = draw_section("Datas", y)
    p.setFont("Helvetica", 11)
    p.drawString(70, y, f"Data da solicitação: {permuta.data_solicitacao.strftime('%d/%m/%Y %H:%M')}")
    y -= 20
    if permuta.data_decisao:
        p.drawString(70, y, f"Data da decisão: {permuta.data_decisao.strftime('%d/%m/%Y %H:%M')}")
        y -= 20
    if permuta.usuario_decisor:
        decisor = permuta.usuario_decisor.get_full_name() or permuta.usuario_decisor.username
        p.drawString(70, y, f"Decidida por: {decisor}")
        y -= 20
    y -= 10

    # Seção: Reposição
    y = draw_section("Reposição", y)
    p.setFont("Helvetica", 11)
    if permuta.tem_reposicao():
        p.drawString(70, y, f"Data da reposição: {permuta.reposicao.data_reposicao.strftime('%d/%m/%Y')}")
        y -= 20
        if permuta.reposicao.observacao:
            p.drawString(70, y, f"Observações: {permuta.reposicao.observacao[:80]}")
            y -= 20
    else:
        p.drawString(70, y, "Não há reposição registrada para esta permuta.")
        y -= 20

    # Seção: Motivo
    y = draw_section("Motivo da Permuta", y)
    p.setFont("Helvetica", 11)
    
    # Quebrar motivo em linhas
    motivo = permuta.motivo or "Não informado"
    palavras = motivo.split()
    linha = ""
    for palavra in palavras:
        if len(linha) + len(palavra) + 1 <= 80:
            linha += (" " if linha else "") + palavra
        else:
            p.drawString(70, y, linha)
            y -= 15
            linha = palavra
    if linha:
        p.drawString(70, y, linha)
        y -= 20

    # Rodapé
    p.setFillColor(colors.HexColor('#6c757d'))
    p.setFont("Helvetica-Oblique", 8)
    p.drawString(50, 50, "Este documento é um comprovante oficial do Sistema de Permuta de Aulas.")
    p.drawRightString(width-50, 50, f"Página 1 de 1")

    p.showPage()
    p.save()

    return response